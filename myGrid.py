from pathlib import Path

import numpy as np
import openpyxl
from rich import print

from Sap2000py import Saproject

ProjectName = "myGrid"

# full path to the model
ModelPath = Path(".\Test\\" + ProjectName + ".sdb")

# Create a Sap2000py obj (default: attatch to instance and create if not exist)
Sap = Saproject()

# Change Sap api settings if you want
# Sap.createSap(AttachToInstance = True, SpecifyPath = False, ProgramPath = "your sap2000 path")

# Open Sap2000 program
Sap.openSap()

# Open Sap2000 sdb file (create if not exist, default: CurrentPath\\NewSapProj.sdb)
Sap.File.Open(ModelPath)
# And print the project information
#Sap.getProjectInfo()

# Can also creat a new blank model (or other models you need like 2DFrame / 3DFrame / etc.)
Sap.File.New_Blank()

# Get the model information (getSapVersion() will print and also returns the version)
vsesion = Sap.SapVersion
Sap.getSapVersion()

# Check your units (getUnits() will print and also returns the units)
unit = Sap.Units
Sap.getUnits()
# set Units as you wish (Just type in Sap.setUnits(""), it will show you the options in your IDE)
Sap.setUnits("KN_m_C")

# Add Materials
#Sap.Scripts.AddCommonMaterialSet(standard = "JTG")

# 
Sap.core.create_grid() # {"NumberBaysX" : 1, "NumberStorys" : 1}
#print("base_points : ", Sap.base_points)
#print("Scenarios : ", Sap.Scenarios)
#print("JointsOfInterest : ", Sap.JointsOfInterest)

#
Sap.RefreshView(0, False)

# Add elements to your group
#Sap.Scripts.Group.RemovefromGroup("base_points", Sap.base_points, "Point")
groups = [
    {"name": "base_points", "type": "Point"},
    #{"name": "girders", "type": "Frame"},
    #{"name": "crossbeams", "type": "Frame"}
]
for g in groups:
    #print("g: ", g['name'], g)
    Sap.Scripts.Group.AddtoGroup(g["name"], getattr(Sap, g["name"]), g["type"])
    items = Sap.Scripts.Group.GetElements(g["name"])
    #print(f'{g["name"]} : {items}')

# Select the group you need
#Sap.Scripts.Group.Select("base_points")

# Save file
Sap.File.Save(ModelPath)

# Remove all cases for analysis
Sap.Scripts.Analyze.RemoveCases("All")
# Select cases for analysis
CaseName = ["DEAD", "MODAL", "G1k", "G2k", "Q1k"] + Sap.Scenarios
Sap.Scripts.Analyze.AddCases(CaseName = CaseName)
# Delete Results
Sap.Scripts.Analyze.DeleteResults("All")
# Run analysis
Sap.Scripts.Analyze.RunAll()

# init a workbook
wb = openpyxl.Workbook()

# select combo for output
xcomboName = "Scenario 01"
comboName = "Scenario 02"
Sap.Scripts.SelectCombo_Case([comboName])

# post processing > xlsx
FileName = Path(".\Test\\" + ProjectName + "_" + comboName + ".xlsx")

# get Joint reaction result by group name
Name, AbsReaction, MaxReaction, MinReaction = Sap.Scripts.GetResults.JointReact_by_Group("base_points")
#print("Name, AbsReaction: ", Name, AbsReaction)
#print("MaxReaction, MinReaction: ", MaxReaction, MinReaction)
ws = wb.create_sheet("base_points-jointforces") # insert at the end (default)
Sap.Scripts.writecell(ws, np.array([["F1", "F2", "F3", "M1", "M2", "M3"]]), "B1" )
Sap.Scripts.writecell(ws, AbsReaction[: , [0, 1, 2, 3, 4, 5] ], "B2" )
for i, v in enumerate(Name):
    ws[f"A{i + 2}"].value = v

# get Joint force result by group name
GroupFrames = [i for i in getattr(Sap, "GroupNames") if i.split('-')[0] == "Frame"]
#print("GroupFrames: ", GroupFrames)
for g in GroupFrames:
    #print(g)
    ret = Sap.Results.Frame.JointForce(g, ItemTypeElm = 2)
    ws = wb.create_sheet(f"{g}-jointforce") # insert at the end (default)
    headers = ["Obj", "Elm", "PointElm", "LoadCase", "StepType", "F1", "F2", "F3", "V2", "T", "M3"]
    #print(len(headers) + 1)
    Sap.Scripts.writecell(ws, np.array([headers]), "A1" )
    Sap.Scripts.writecell(ws, np.transpose(np.array(ret[1 : len(headers) + 1])), "A2" )

"""
# get Joint force result by group name
GroupFrames = [i for i in getattr(Sap, "GroupNames") if i.split('-')[0] == "Frame"]
#print("GroupFrames: ", GroupFrames)
for g in GroupFrames:
    #print(g)
    ret = Sap.Results.Frame.Force(g, ItemTypeElm = 2)
    ws = wb.create_sheet(f"{g}-frameforce") # insert at the end (default)
    headers = ["Obj", "Elm", "PointElm", "LoadCase", "StepType", "P", "V2", "V3", "T", "M2", "M3"]
    #print(len(headers) + 1)
    Sap.Scripts.writecell(ws, np.array([headers]), "A1" )
    Sap.Scripts.writecell(ws, np.transpose(np.array(ret[1 : len(headers) + 1])), "A2" )
"""
    
# get joint displ result by group name
GroupPoints = [i for i in getattr(Sap, "GroupNames") if i.split('-')[0] == "Point"]
#print("GroupPoints: ", GroupPoints)
for g in GroupPoints:
    #print(g)
    ret = Sap.Results.Joint.Displ(g, ItemTypeElm = 2)
    ws = wb.create_sheet(f"{g}-displ") # insert at the end (default)
    headers = ["Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2", "R3"]
    #print(len(headers) + 1)
    Sap.Scripts.writecell(ws, np.array([headers]), "A1" )
    Sap.Scripts.writecell(ws, np.transpose(np.array(ret[1 : len(headers) + 1])), "A2" )

wb.save(FileName)

# Save your file with a Filename(default: your ModelPath)
#Sap.File.Save()
Sap.File.Save(ModelPath)

# Don't forget to close the program
#Sap.closeSap()