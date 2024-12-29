from pathlib import Path

import numpy as np
import openpyxl
from rich import print

from Sap2000py import Saproject

ProjectName = "myTest"

# full path to the model
ModelPath = Path(".\Test\\" + ProjectName + ".sdb")

# Create a Sap2000py obj (default: attatch to instance and create if not exist)
Sap = Saproject()

# Change Sap api settings if you want
# Sap.createSap(AttachToInstance = True, SpecifyPath = False, ProgramPath = "your sap2000 path")

# Open Sap2000 program
Sap.openSap()

# Open Sap2000 sdb file (create if not exist, default: CurrentPath\\NewSapProj.sdb)
Sap.File.Open(ModelPath)
# And print the project information
#Sap.getProjectInfo()

# Can also creat a new blank model (or other models you need like 2DFrame / 3DFrame / etc.)
Sap.File.New_Blank()

# Get the model information (getSapVersion() will print and also returns the version)
vsesion = Sap.SapVersion
Sap.getSapVersion()

# Check your units (getUnits() will print and also returns the units)
unit = Sap.Units
Sap.getUnits()
# set Units as you wish (Just type in Sap.setUnits(""), it will show you the options in your IDE)
Sap.setUnits("KN_m_C")

# Add China Common Material Set·
#Sap.Scripts.AddCommonMaterialSet(standard = "JTG")

# add cartesian joints
#joint_coordinates = np.array([ [0,0,0], [0,0,3], [6,0,3], [6,0,0]])
#Sap.core.add_cartesian_joints(joint_coordinates)
# You can also Add Joints once a time : Sap.Assign.PointObj.AddCartesian(x=0, y=0, z=0)
# After using this script to add joints, you can see all the joints in var Sap.coord_joints
#print("joint_coordinates : ", Sap.joint_coordinates)

# add frame elements
#Sap.core.add_frame_by_points([[1,2], [2,3], [3,4]])
#print("connectivity_frame : ", Sap.connectivity_frame)

# 
Sap.core.create_3d_frame({"NumberBaysX" : 1, "NumberStorys" : 1})
#print("base_points : ", Sap.base_points)
#print("columns : ", Sap.columns)
#print("beams_x : ", Sap.beams_x)
#print("beams_y : ", Sap.beams_y)

#
"""
Sap.File.New_2DFrame(
                    TempType = "EccentricBraced", # "PortalFrame","ConcentricBraced","EccentricBraced"
                    NumberStorys = 3, StoryHeight = 3, NumberBays = 6, BayWidth = 6,
                    Restraint = True, Beam="Default", Column="Default", Brace="Default")

Sap.File.New_3DFrame(
                    TempType = 0, # OpenFrame = 0, PerimeterFrame = 1, BeamSlab = 2, FlatPlate = 3
                    NumberStorys = 3, StoryHeight = 3, NumberBayX = 4, BayWidthX = 4, NumberBaysY = 3, BayWidthY = 3,
                    Restraint = True, Beam = "Default", Column = "Default", Area = "Default",
                    NumberXDivisions = 4, NumberYDivisions = 4)
"""
#
#Sap.File.New_Wall(NumberXDivisions = 10, DivisionWidthX = 1, NumberZDivisions = 10, DivisionWidthZ = 1, Restraint = True, Area = "Default")

#
#Sap.File.New_SolidBlock(XWidth = 4, YWidth = 3, Height = 3, Restraint = True, Solid = "Default", NumberXDivisions = 4, NumberYDivisions = 3, NumberZDivisions = 3)

#
Sap.RefreshView(0, False)

# Add elements to your group
#Sap.Scripts.Group.RemovefromGroup("base_points", Sap.base_points, "Point")
groups = [
    {"name": "base_points", "type": "Point"},
    {"name": "columns", "type": "Frame"},
    {"name": "beams_x", "type": "Frame"},
    {"name": "beams_y", "type": "Frame"}
]
for g in groups:
    #print("g: ", g['name'], g)
    Sap.Scripts.Group.AddtoGroup(g["name"], getattr(Sap, g["name"]), g["type"])
    items = Sap.Scripts.Group.GetElements(g["name"])
    #print(f'{g["name"]} : {items}')

# Select the group you need
#Sap.Scripts.Group.Select("base_points")

# Save file
Sap.File.Save(ModelPath)

# Remove all cases for analysis
Sap.Scripts.Analyze.RemoveCases("All")
# Select cases for analysis
Sap.Scripts.Analyze.AddCases(CaseName = ["DEAD", "MODAL", "G1k", "G2k", "Q1k"])
# Delete Results
Sap.Scripts.Analyze.DeleteResults("All")
# Run analysis
Sap.Scripts.Analyze.RunAll()

# post processing > xlsx
FileName = Path(".\Test\\" + ProjectName + ".xlsx")

# open excel
wb = openpyxl.Workbook()

# select combo for output
comboName = "Fd"
Sap.Scripts.SelectCombo_Case([comboName])

# get Joint reaction result by group name
Name, AbsReaction, MaxReaction, MinReaction = Sap.Scripts.GetResults.JointReact_by_Group("base_points")
#print("Name, AbsReaction: ", Name, AbsReaction)
#print("MaxReaction, MinReaction: ", MaxReaction, MinReaction)
ws = wb.create_sheet(f"{comboName}-base_points-jointforces") # insert at the end (default)
Sap.Scripts.writecell(ws, np.array([["F1", "F2", "F3", "M1", "M2", "M3"]]), "B1" )
Sap.Scripts.writecell(ws, AbsReaction[: , [0, 1, 2, 3, 4, 5] ], "B2" )
for i, v in enumerate(Name):
    ws[f"A{i + 2}"].value = v

frames = [
    {"name": "columns", "type": "Frame"},
    {"name": "beams_x", "type": "Frame"},
    {"name": "beams_y", "type": "Frame"}
]
for g in frames:
    name = g["name"]

    # get Frame force result by group name
    #Name, AbsReaction, MaxReaction, MinReaction = Sap.Scripts.GetResults.ElementJointForce_by_Group(name)
    #print("Name, AbsReaction: ", Name, AbsReaction)
    #print("MaxReaction, MinReaction: ", MaxReaction, MinReaction)
    """
    ws = wb.create_sheet(f"{comboName}-{name}-jointforces") # insert at the end (default)
    Sap.Scripts.writecell(ws, np.array([["F1", "F2", "F3", "M1", "M2", "M3"]]), "B1" )
    Sap.Scripts.writecell(ws, AbsReaction[: , [0, 1, 2, 3, 4, 5] ], "B2" )
    for i, v in enumerate(Name):
        ws[f"A{i + 2}"].value = v
    """

    # get Frame force result by group name
    Name, AbsReaction, MaxReaction, MinReaction = Sap.Scripts.GetResults.ElementForce_by_Group(name)
    #print("Name, AbsReaction: ", Name, AbsReaction)
    #print("MaxReaction, MinReaction: ", MaxReaction, MinReaction)
    ws = wb.create_sheet(f"{comboName}-{name}-frameforces") # insert at the end (default)
    Sap.Scripts.writecell(ws, np.array([["P", "V2", "V3", "T", "M2", "M3"]]), "B1" )
    Sap.Scripts.writecell(ws, AbsReaction[: , [0, 1, 2, 3, 4, 5] ], "B2" )
    for i, v in enumerate(Name):
        ws[f"A{i + 2}"].value = v

# get reactions under earthquake time history
# select combo for output
#Sap.Scripts.SelectCombo_Case(["E2YEarthquake"])

# get Frame force result by group name
#Name,EleAbsForce,EleMaxForce,EleMinForce = Sap.Scripts.GetResults.ElementJointForce_by_Group("PierBottom")
# write in excel (here we need F3,F1,M2 --> [2,0,4]),,"D30" is the top left corner of the matrix
#Sap.Scripts.writecell(ws,EleAbsForce[:,[2,0,4]],"D30")

wb.save(FileName)

# Save your file with a Filename(default: your ModelPath)
#Sap.File.Save()
Sap.File.Save(ModelPath)

# Don't forget to close the program
#Sap.closeSap()